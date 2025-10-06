#!/usr/bin/env python3
"""
Create a simple test XLSX file for testing the Go XLSX parser.
"""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not available. Install with 'pip install openpyxl'")
    sys.exit(1)

def create_test_xlsx():
    """Create a simple test XLSX file."""

    # Create a new workbook
    wb = openpyxl.Workbook()

    # Get the default sheet and rename it
    sheet1 = wb.active
    sheet1.title = "Employees"

    # Add headers with formatting
    headers = ["ID", "Name", "Department", "Salary", "Email", "Website"]
    for col, header in enumerate(headers, 1):
        cell = sheet1.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Add some sample data
    data = [
        [1, "John Doe", "Engineering", 75000, "john.doe@company.com", "https://johndoe.dev"],
        [2, "Jane Smith", "Marketing", 65000, "jane.smith@company.com", "https://janesmith.com"],
        [3, "Bob Wilson", "Sales", 55000, "bob.wilson@company.com", "https://bobwilson.net"],
        [4, "Alice Brown", "Engineering", 80000, "alice.brown@company.com", "https://alicebrown.tech"],
        [5, "Charlie Davis", "HR", 60000, "charlie.davis@company.com", ""],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet1.cell(row=row_idx, column=col_idx, value=value)

    # Add a second sheet with a simple table
    sheet2 = wb.create_sheet("Departments")
    dept_headers = ["Department", "Manager", "Budget"]
    for col, header in enumerate(dept_headers, 1):
        cell = sheet2.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        cell.fill = openpyxl.styles.PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")

    dept_data = [
        ["Engineering", "Sarah Johnson", 500000],
        ["Marketing", "Mike Rodriguez", 200000],
        ["Sales", "Linda Thompson", 300000],
        ["HR", "David Lee", 150000],
    ]

    for row_idx, row_data in enumerate(dept_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet2.cell(row=row_idx, column=col_idx, value=value)

    # Add a comment to one of the cells
    sheet1.cell(row=2, column=2).comment = openpyxl.comments.Comment(
        "This is John's record", "System"
    )

    # Add a formula
    sheet2.cell(row=6, column=2, value="Total Budget:")
    sheet2.cell(row=6, column=3, value="=SUM(C2:C5)")

    # Save the workbook
    test_file = Path("tests/assets/test_sample.xlsx")
    test_file.parent.mkdir(parents=True, exist_ok=True)

    wb.save(test_file)
    print(f"Created test XLSX file: {test_file}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Employees sheet: {len(data)+1} rows, {len(headers)} columns")
    print(f"Departments sheet: {len(dept_data)+1} rows, {len(dept_headers)} columns")
    return test_file

if __name__ == "__main__":
    create_test_xlsx()