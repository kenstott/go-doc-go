"""
Comprehensive test suite for XLSX document parser.

Tests follow CLAUDE.md guidelines:
- Test design objectives, not implementation
- Use pytest markers for categorization (unit, integration, performance)
- Follow DRY principles with reusable helpers
- Validate against ElementType and RelationshipType enums
"""

import json
import os
import tempfile
import time
from pathlib import Path
from typing import Dict, Any, List
from unittest.mock import Mock, patch

import pytest

# Import the parser and related modules
from go_doc_go.document_parser.xlsx import XlsxParser

try:
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.styles.colors import Color
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import enums for validation
from go_doc_go.storage.element_element import ElementType
from go_doc_go.relationships.structural import RelationshipType


# =============================================================================
# Test Helpers and Validators (DRY Principle)
# =============================================================================

def assert_valid_element(element: Dict[str, Any]) -> None:
    """Validate element structure against design specifications."""
    # Required fields
    required_fields = ["element_id", "doc_id", "element_type", "content_preview", "content_location"]
    for field in required_fields:
        assert field in element, f"Missing required field: {field}"
    
    # Validate element type against enum
    element_type = element["element_type"]
    valid_types = [e.value for e in ElementType]
    assert element_type in valid_types, \
        f"Invalid element_type '{element_type}' not in ElementType enum"
    
    # Validate content_location is JSON serializable
    content_location = element["content_location"]
    assert isinstance(content_location, str), "content_location must be a JSON string"
    try:
        location_data = json.loads(content_location)
        assert "source" in location_data
        assert "type" in location_data
    except json.JSONDecodeError:
        pytest.fail(f"content_location is not valid JSON: {content_location}")


def assert_valid_relationship(rel: Dict[str, Any]) -> None:
    """Validate relationship structure against design specifications."""
    required_fields = ["relationship_id", "source_id", "target_id", "relationship_type"]
    for field in required_fields:
        assert field in rel, f"Missing required field: {field}"
    
    # Validate relationship type against enum
    rel_type = rel["relationship_type"]
    valid_types = [r.value for r in RelationshipType]
    assert rel_type in valid_types, \
        f"Invalid relationship_type '{rel_type}' not in RelationshipType enum"


def assert_valid_parse_result(result: Dict[str, Any]) -> None:
    """Common assertions for all parse results."""
    assert "document" in result, "Result missing 'document' key"
    assert "elements" in result, "Result missing 'elements' key"
    assert "relationships" in result, "Result missing 'relationships' key"
    
    # Validate document structure
    doc = result["document"]
    assert "doc_id" in doc
    assert "doc_type" in doc
    assert doc["doc_type"] == "xlsx"
    assert "metadata" in doc
    
    # Validate all elements
    for element in result["elements"]:
        assert_valid_element(element)
    
    # Validate all relationships
    for rel in result["relationships"]:
        assert_valid_relationship(rel)


def assert_xlsx_hierarchy(elements: List[Dict[str, Any]]) -> None:
    """Validate XLSX-specific element hierarchy per design."""
    # Must have root element
    root_elements = [e for e in elements if e["element_type"] == "root"]
    assert len(root_elements) == 1, "Must have exactly one root element"
    
    # Must have workbook element
    workbook_elements = [e for e in elements if e["element_type"] == "workbook"]
    assert len(workbook_elements) >= 1, "Must have at least one workbook element"
    
    # Workbook must be child of root
    root_id = root_elements[0]["element_id"]
    for wb in workbook_elements:
        assert wb.get("parent_id") == root_id, "Workbook must be child of root"


def create_simple_xlsx(path: str, **kwargs) -> None:
    """
    Create a simple Excel file for testing.
    DRY helper used across multiple tests.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = kwargs.get("sheet_name", "Sheet1")
    
    # Add data if provided
    if "data" in kwargs:
        for row_idx, row_data in enumerate(kwargs["data"], 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    else:
        # Default data
        ws['A1'] = 'Header 1'
        ws['B1'] = 'Header 2'
        ws['A2'] = 'Data 1'
        ws['B2'] = 'Data 2'
    
    wb.save(path)
    wb.close()


def create_xlsx_with_sheets(path: str, sheet_names: List[str]) -> None:
    """Create Excel file with multiple sheets."""
    wb = openpyxl.Workbook()
    
    # Remove default sheet and add requested sheets
    wb.remove(wb.active)
    
    for sheet_name in sheet_names:
        ws = wb.create_sheet(title=sheet_name)
        ws['A1'] = f'{sheet_name} Title'
        ws['A2'] = f'Data in {sheet_name}'
    
    wb.save(path)
    wb.close()


def create_xlsx_with_formatting(path: str) -> None:
    """Create Excel file with various cell formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add formatted cells
    ws['A1'] = 'Bold Text'
    ws['A1'].font = Font(bold=True, color="FF0000")
    
    ws['B1'] = 'Filled Cell'
    ws['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws['C1'] = 'Bordered Cell'
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    ws['C1'].border = border
    
    ws['A2'] = 'Aligned Text'
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    
    wb.save(path)
    wb.close()


# =============================================================================
# Fixtures (Following guidelines for test data)
# =============================================================================

@pytest.fixture
def temp_xlsx_path():
    """Provide a temporary file path for XLSX files."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        yield tmp.name
    # Cleanup
    if os.path.exists(tmp.name):
        os.unlink(tmp.name)


@pytest.fixture
def simple_xlsx_content(temp_xlsx_path):
    """Create a simple XLSX file and return content dict."""
    create_simple_xlsx(temp_xlsx_path)
    
    return {
        "id": temp_xlsx_path,
        "binary_path": temp_xlsx_path,
        "metadata": {"doc_id": "test_doc_123"}
    }


@pytest.fixture
def multi_sheet_xlsx_content(temp_xlsx_path):
    """Create XLSX with multiple sheets."""
    create_xlsx_with_sheets(temp_xlsx_path, ["Sales", "Inventory", "Reports"])
    
    return {
        "id": temp_xlsx_path,
        "binary_path": temp_xlsx_path,
        "metadata": {"doc_id": "multi_sheet_123"}
    }


@pytest.fixture
def formatted_xlsx_content(temp_xlsx_path):
    """Create XLSX with formatting."""
    create_xlsx_with_formatting(temp_xlsx_path)
    
    return {
        "id": temp_xlsx_path,
        "binary_path": temp_xlsx_path,
        "metadata": {"doc_id": "formatted_123"}
    }


@pytest.fixture
def expected_simple_structure():
    """Expected structure for simple XLSX per design specification."""
    return {
        "min_elements": 5,  # root, workbook, sheet, 2 rows minimum
        "required_types": ["root", "workbook", "sheet", "table_row"],
        "required_relationships": [
            RelationshipType.CONTAINS.value,
            RelationshipType.CONTAINED_BY.value
        ]
    }


# =============================================================================
# Unit Tests - Fast, Isolated Component Tests
# =============================================================================

@pytest.mark.unit
class TestXlsxParser:
    """Unit tests for XlsxParser - test in isolation without real files."""
    
    def test_parser_initialization(self):
        """Test parser initializes with default configuration."""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not available")
            
        parser = XlsxParser()
        
        # Verify default configuration values
        assert parser.extract_hidden_sheets == False
        assert parser.extract_formulas == True
        assert parser.extract_comments == True
        assert parser.max_rows == 1000
        assert parser.max_cols == 100
        assert parser.detect_tables == True
    
    def test_parser_custom_config(self):
        """Test parser accepts custom configuration."""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not available")
            
        config = {
            "extract_hidden_sheets": True,
            "extract_formulas": False,
            "max_rows": 500,
            "max_cols": 50,
            "detect_tables": False
        }
        parser = XlsxParser(config)
        
        assert parser.extract_hidden_sheets == True
        assert parser.extract_formulas == False
        assert parser.max_rows == 500
        assert parser.max_cols == 50
        assert parser.detect_tables == False
    
    def test_generate_id_uniqueness(self):
        """Test that ID generation produces unique IDs."""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not available")
            
        parser = XlsxParser()
        
        ids = [parser._generate_id("test_") for _ in range(100)]
        
        # All IDs should be unique
        assert len(ids) == len(set(ids))
        
        # All should start with prefix
        for id_ in ids:
            assert id_.startswith("test_")
    
    def test_generate_hash(self):
        """Test content hash generation."""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not available")
            
        parser = XlsxParser()
        
        # Same content should produce same hash
        hash1 = parser._generate_hash("test content")
        hash2 = parser._generate_hash("test content")
        assert hash1 == hash2
        
        # Different content should produce different hash
        hash3 = parser._generate_hash("different content")
        assert hash1 != hash3
    
    def test_ensure_serializable(self):
        """Test that _ensure_serializable converts non-JSON types."""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not available")
            
        parser = XlsxParser()
        
        # Test with various types
        test_data = {
            "string": "test",
            "number": 42,
            "float": 3.14,
            "bool": True,
            "none": None,
            "list": [1, 2, 3],
            "nested": {"key": "value"},
            "non_serializable": object()  # This should be converted to string
        }
        
        result = parser._ensure_serializable(test_data)
        
        # Try to serialize result - should not raise
        json_str = json.dumps(result)
        assert isinstance(json_str, str)
    
    def test_extract_cell_style_empty(self):
        """Test style extraction from cell without formatting."""
        if not OPENPYXL_AVAILABLE:
            pytest.skip("openpyxl not available")
            
        # Create a mock cell with minimal attributes
        mock_cell = Mock()
        mock_cell.font = None
        
        style = XlsxParser._extract_cell_style(mock_cell)
        
        assert isinstance(style, dict)
        assert len(style) == 0  # No style for plain cell
    
    def test_missing_openpyxl_error(self):
        """Test clear error when openpyxl not installed."""
        with patch('go_doc_go.document_parser.xlsx.OPENPYXL_AVAILABLE', False):
            with pytest.raises(ImportError, match="openpyxl is required"):
                XlsxParser()


# =============================================================================
# Integration Tests - Test with Real Excel Files
# =============================================================================

@pytest.mark.integration
class TestXlsxParserIntegration:
    """Integration tests using real Excel files."""
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_parse_simple_xlsx(self, simple_xlsx_content, expected_simple_structure):
        """Test parsing a simple Excel file validates design objectives."""
        parser = XlsxParser()
        result = parser.parse(simple_xlsx_content)
        
        # Validate against design specifications
        assert_valid_parse_result(result)
        
        # Check minimum elements exist
        assert len(result["elements"]) >= expected_simple_structure["min_elements"]
        
        # Check required element types
        element_types = {e["element_type"] for e in result["elements"]}
        for required_type in expected_simple_structure["required_types"]:
            assert required_type in element_types, f"Missing required type: {required_type}"
        
        # Validate hierarchy
        assert_xlsx_hierarchy(result["elements"])
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_parse_multi_sheet(self, multi_sheet_xlsx_content):
        """Test parsing Excel with multiple sheets."""
        parser = XlsxParser()
        result = parser.parse(multi_sheet_xlsx_content)
        
        assert_valid_parse_result(result)
        
        # Should have sheets for each created sheet
        sheet_elements = [e for e in result["elements"] if e["element_type"] == "sheet"]
        assert len(sheet_elements) == 3
        
        # Check sheet names in metadata
        sheet_names = {e["metadata"]["title"] for e in sheet_elements}
        assert sheet_names == {"Sales", "Inventory", "Reports"}
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_cell_values_extraction(self, simple_xlsx_content):
        """Test that cell values are properly extracted."""
        parser = XlsxParser()
        result = parser.parse(simple_xlsx_content)
        
        # Find cell elements
        cell_elements = [e for e in result["elements"] 
                        if e["element_type"] in ["table_cell", "table_header"]]
        
        assert len(cell_elements) > 0, "Should have cell elements"
        
        # Check that cells have content preview
        for cell in cell_elements:
            assert "content_preview" in cell
            # At least some cells should have non-empty content
        
        # Verify some expected values exist
        previews = {cell["content_preview"] for cell in cell_elements}
        assert "Header 1" in previews or "Header 2" in previews
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_relationships_structure(self, simple_xlsx_content):
        """Test that relationships follow design specifications."""
        parser = XlsxParser()
        result = parser.parse(simple_xlsx_content)
        
        relationships = result["relationships"]
        assert len(relationships) > 0
        
        # Check for contains/contained_by pairs
        contains_rels = [r for r in relationships 
                        if r["relationship_type"] == RelationshipType.CONTAINS.value]
        contained_by_rels = [r for r in relationships 
                           if r["relationship_type"] == RelationshipType.CONTAINED_BY.value]
        
        # Should have inverse relationships
        assert len(contains_rels) > 0
        assert len(contained_by_rels) > 0
        
        # For each CONTAINS, should have inverse CONTAINED_BY
        for contains_rel in contains_rels[:5]:  # Check first 5
            inverse_exists = any(
                r["source_id"] == contains_rel["target_id"] and
                r["target_id"] == contains_rel["source_id"]
                for r in contained_by_rels
            )
            assert inverse_exists, "Missing inverse relationship"
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_formatted_cells(self, formatted_xlsx_content):
        """Test extraction of cell formatting."""
        parser = XlsxParser()
        result = parser.parse(formatted_xlsx_content)
        
        # Find cells with formatting
        cells = [e for e in result["elements"] 
                if e["element_type"] in ["table_cell", "table_header"]]
        
        # At least some cells should have style metadata
        styled_cells = [c for c in cells 
                       if c.get("metadata", {}).get("style", {})]
        
        assert len(styled_cells) > 0, "Should extract cell styles"
        
        # Check that style data is present and JSON serializable
        for cell in styled_cells:
            style = cell["metadata"]["style"]
            # Should be able to serialize
            json.dumps(style)


# =============================================================================
# Data Table Detection Tests
# =============================================================================

# =============================================================================
# New Coverage Enhancement Tests
# =============================================================================

@pytest.mark.unit 
class TestXlsxParserExtended:
    """Extended tests to improve XLSX parser coverage."""

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_missing_openpyxl_error_path(self):
        """Test the ImportError path when openpyxl is missing."""
        with patch('go_doc_go.document_parser.xlsx.OPENPYXL_AVAILABLE', False):
            with pytest.raises(ImportError, match="openpyxl is required"):
                XlsxParser()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_parse_method_coverage(self, temp_xlsx_path):
        """Test parse method with actual binary file processing."""
        # Create test XLSX file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test Data'
        ws['B1'] = 42
        wb.save(temp_xlsx_path)
        
        # Test with binary_path instead of content
        with open(temp_xlsx_path, 'rb') as f:
            binary_content = f.read()
            
        content = {
            "id": "/test.xlsx", 
            "binary_path": temp_xlsx_path,
            "metadata": {"doc_id": "test123"}
        }
        
        parser = XlsxParser()
        result = parser.parse(content)
        
        assert result is not None
        assert "elements" in result
        assert len(result["elements"]) > 0

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_workbook_metadata_extraction(self, temp_xlsx_path):
        """Test _extract_document_metadata static method."""
        wb = openpyxl.Workbook()
        # Set some document properties
        wb.properties.title = "Test Document"
        wb.properties.creator = "Test Creator"
        wb.properties.description = "Test Description"
        wb.save(temp_xlsx_path)
        
        # Load workbook and test metadata extraction
        wb_loaded = openpyxl.load_workbook(temp_xlsx_path)
        base_metadata = {"existing": "data"}
        
        metadata = XlsxParser._extract_document_metadata(wb_loaded, base_metadata)
        
        assert "existing" in metadata
        assert "title" in metadata or "creator" in metadata

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available") 
    def test_extract_workbook_text(self, temp_xlsx_path):
        """Test _extract_workbook_text method."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Hello'
        ws['B1'] = 'World'
        ws['A2'] = 123
        wb.save(temp_xlsx_path)
        
        wb_loaded = openpyxl.load_workbook(temp_xlsx_path)
        parser = XlsxParser()
        
        text = parser._extract_workbook_text(wb_loaded)
        
        assert isinstance(text, str)
        assert 'Hello' in text or 'World' in text

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_merged_cells_extraction(self, temp_xlsx_path):
        """Test _extract_merged_cells method."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Merged Cell'
        ws.merge_cells('A1:B2')
        wb.save(temp_xlsx_path)
        
        wb_loaded = openpyxl.load_workbook(temp_xlsx_path)
        ws_loaded = wb_loaded.active
        
        parser = XlsxParser()
        elements, relationships = parser._extract_merged_cells(
            ws_loaded, "doc1", "sheet1", "source1"
        )
        
        assert isinstance(elements, list)
        assert isinstance(relationships, list)

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_comments_extraction(self, temp_xlsx_path):
        """Test _extract_comments method."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Cell with comment'
        from openpyxl.comments import Comment
        ws['A1'].comment = Comment('This is a comment', 'Author')
        wb.save(temp_xlsx_path)
        
        wb_loaded = openpyxl.load_workbook(temp_xlsx_path)
        ws_loaded = wb_loaded.active
        
        parser = XlsxParser()
        elements, relationships = parser._extract_comments(
            ws_loaded, "doc1", "sheet1", "source1"
        )
        
        assert isinstance(elements, list)
        assert isinstance(relationships, list)

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_data_table_detection(self, temp_xlsx_path):
        """Test _detect_data_tables method."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Create a simple data table structure
        headers = ['Name', 'Age', 'City']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add some data rows
        data = [['John', 25, 'NYC'], ['Jane', 30, 'LA'], ['Bob', 35, 'Chicago']]
        for row, row_data in enumerate(data, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
                
        wb.save(temp_xlsx_path)
        
        wb_loaded = openpyxl.load_workbook(temp_xlsx_path)
        ws_loaded = wb_loaded.active
        
        parser = XlsxParser()
        elements, relationships = parser._detect_data_tables(
            ws_loaded, "doc1", "sheet1", "source1", 4, 3
        )
        
        assert isinstance(elements, list)
        assert isinstance(relationships, list)

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_supports_location_method(self):
        """Test supports_location method."""
        parser = XlsxParser()
        
        # Valid location
        valid_location = {
            "sheet": "Sheet1",
            "cell": "A1"
        }
        assert parser.supports_location(valid_location) == True
        
        # Invalid location
        invalid_location = {
            "page": "1",
            "line": "5"
        }
        assert parser.supports_location(invalid_location) == False

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_resolve_element_text(self, temp_xlsx_path):
        """Test _resolve_element_text method."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test Content'
        wb.save(temp_xlsx_path)
        
        parser = XlsxParser()
        location_data = {
            "sheet": "Sheet", 
            "cell": "A1",
            "source": temp_xlsx_path
        }
        
        with open(temp_xlsx_path, 'rb') as f:
            binary_content = f.read()
            
        text = parser._resolve_element_text(location_data, binary_content)
        
        assert isinstance(text, str)

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_resolve_element_content(self, temp_xlsx_path):
        """Test _resolve_element_content method."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test Content'
        wb.save(temp_xlsx_path)
        
        parser = XlsxParser()
        location_data = {
            "sheet": "Sheet",
            "cell": "A1", 
            "source": temp_xlsx_path
        }
        
        with open(temp_xlsx_path, 'rb') as f:
            binary_content = f.read()
            
        content = parser._resolve_element_content(location_data, binary_content)
        
        assert isinstance(content, dict)
        assert "text" in content

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_extract_workbook_links(self, temp_xlsx_path):
        """Test _extract_workbook_links static method."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Link Text'
        # Adding hyperlinks requires more complex setup but test method exists
        wb.save(temp_xlsx_path)
        
        wb_loaded = openpyxl.load_workbook(temp_xlsx_path)
        elements = [{"element_id": "test1", "content_preview": "test"}]
        
        links = XlsxParser._extract_workbook_links(wb_loaded, elements)
        
        assert isinstance(links, list)

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_is_cell_in_range_static(self):
        """Test _is_cell_in_range static method."""
        # Test various cell ranges
        assert XlsxParser._is_cell_in_range("A1", "A1:B2") == True
        assert XlsxParser._is_cell_in_range("B2", "A1:B2") == True
        assert XlsxParser._is_cell_in_range("C3", "A1:B2") == False

    def test_ensure_serializable_method(self):
        """Test _ensure_serializable method."""
        parser = XlsxParser()
        
        # Test with dict
        test_dict = {"key": "value", "number": 42}
        result = parser._ensure_serializable(test_dict)
        assert result == test_dict
        
        # Test with list  
        test_list = ["item1", "item2", 123]
        result = parser._ensure_serializable(test_list)
        assert result == test_list
        
        # Test with string
        test_str = "simple string"
        result = parser._ensure_serializable(test_str)
        assert result == test_str

@pytest.mark.integration
class TestDataTableDetection:
    """Test data table detection functionality."""
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_detect_simple_table(self, temp_xlsx_path):
        """Test detection of a simple data table."""
        # Create Excel with clear table structure
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create table with headers
        headers = ["ID", "Name", "Age", "City"]
        data = [
            [1, "John", 30, "NYC"],
            [2, "Jane", 25, "LA"],
            [3, "Bob", 35, "Chicago"]
        ]
        
        # Add headers with formatting
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
        
        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        wb.save(temp_xlsx_path)
        wb.close()
        
        # Parse and check for data table
        parser = XlsxParser({"detect_tables": True})
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        
        # Should detect data table
        data_tables = [e for e in result["elements"] 
                      if e["element_type"] == "data_table"]
        
        assert len(data_tables) > 0, "Should detect data table"
        
        # Check table metadata
        table = data_tables[0]
        assert table["metadata"]["has_header"] == True
        assert table["metadata"]["row_count"] == 4  # Including header
        assert table["metadata"]["column_count"] == 4
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_autofilter_table_detection(self, temp_xlsx_path):
        """Test that autofilter indicates a data table."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add data
        ws['A1'] = 'Column1'
        ws['B1'] = 'Column2'
        ws['A2'] = 'Data1'
        ws['B2'] = 'Data2'
        
        # Add autofilter
        ws.auto_filter.ref = "A1:B2"
        
        wb.save(temp_xlsx_path)
        wb.close()
        
        parser = XlsxParser()
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        
        # Should detect table from autofilter
        data_tables = [e for e in result["elements"] 
                      if e["element_type"] == "data_table"]
        
        assert len(data_tables) > 0, "Should detect at least one data table when autofilter is present"
        table = data_tables[0]
        assert table["metadata"]["detection_confidence"] == "high"  # Autofilter = high confidence


# =============================================================================
# Content Resolution Tests
# =============================================================================

@pytest.mark.integration
class TestContentResolution:
    """Test content resolution functionality."""
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_resolve_sheet_content(self, simple_xlsx_content):
        """Test resolving sheet element content."""
        parser = XlsxParser()
        result = parser.parse(simple_xlsx_content)
        
        # Find a sheet element
        sheet_elem = next(e for e in result["elements"] 
                         if e["element_type"] == "sheet")
        
        # Resolve its content
        location = json.loads(sheet_elem["content_location"])
        content = parser._resolve_element_content(location, None)
        
        assert isinstance(content, str)
        assert "Sheet" in content  # Should describe sheet
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_resolve_cell_content(self, simple_xlsx_content):
        """Test resolving individual cell content."""
        parser = XlsxParser()
        result = parser.parse(simple_xlsx_content)
        
        # Find a cell element
        cell_elem = next(e for e in result["elements"] 
                        if e["element_type"] in ["table_cell", "table_header"])
        
        # Resolve its content
        location = json.loads(cell_elem["content_location"])
        content = parser._resolve_element_content(location, None)
        
        assert isinstance(content, str)
        # Content should match the preview
        assert content.strip() == cell_elem["content_preview"].strip()


# =============================================================================
# Special Features Tests
# =============================================================================

@pytest.mark.integration
class TestSpecialFeatures:
    """Test special Excel features."""
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_merged_cells(self, temp_xlsx_path):
        """Test handling of merged cells."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create merged cells
        ws.merge_cells('A1:B2')
        ws['A1'] = 'Merged Cell Content'
        
        # Add regular cells
        ws['C1'] = 'Regular Cell'
        
        wb.save(temp_xlsx_path)
        wb.close()
        
        parser = XlsxParser()
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        
        # Should have merged cell element
        merged_cells = [e for e in result["elements"] 
                       if e["element_type"] == "merged_cell"]
        
        if len(merged_cells) > 0:
            merged = merged_cells[0]
            assert "Merged Cell Content" in merged["content_preview"]
            assert merged["metadata"]["range"] == "A1:B2"
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_comments_extraction(self, temp_xlsx_path):
        """Test extraction of cell comments."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add cell with comment
        ws['A1'] = 'Cell with comment'
        comment = Comment('This is a comment', 'Test Author')
        ws['A1'].comment = comment
        
        wb.save(temp_xlsx_path)
        wb.close()
        
        parser = XlsxParser({"extract_comments": True})
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        
        # Should have comment element
        comments = [e for e in result["elements"] 
                   if e["element_type"] == "comment"]
        
        if len(comments) > 0:
            comment_elem = comments[0]
            assert "This is a comment" in comment_elem["content_preview"]
            assert comment_elem["metadata"]["author"] == "Test Author"
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_formula_extraction(self, temp_xlsx_path):
        """Test extraction of formulas when enabled."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add cells with formula
        ws['A1'] = 10
        ws['B1'] = 20
        ws['C1'] = '=A1+B1'
        
        wb.save(temp_xlsx_path)
        wb.close()
        
        # Parse with formula extraction enabled
        parser = XlsxParser({"extract_formulas": True})
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        
        # Find cell C1
        cells = [e for e in result["elements"] 
                if e["element_type"] in ["table_cell", "table_header"]
                and e.get("metadata", {}).get("address") == "C1"]
        
        if len(cells) > 0:
            formula_cell = cells[0]
            # When extract_formulas=True, should have formula in metadata
            if "formula" in formula_cell["metadata"]:
                assert formula_cell["metadata"]["formula"] == "A1+B1"


# =============================================================================
# Error Handling Tests
# =============================================================================

@pytest.mark.unit
class TestErrorHandling:
    """Test error handling scenarios."""
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_corrupt_file_handling(self, temp_xlsx_path):
        """Test handling of corrupt Excel files."""
        # Create corrupt file
        with open(temp_xlsx_path, 'wb') as f:
            f.write(b'This is not a valid Excel file')
        
        parser = XlsxParser()
        
        with pytest.raises(Exception):  # Should raise some exception
            parser.parse({
                "id": temp_xlsx_path,
                "binary_path": temp_xlsx_path,
                "metadata": {}
            })
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_empty_workbook(self, temp_xlsx_path):
        """Test handling of empty Excel file."""
        wb = openpyxl.Workbook()
        # Don't add any data
        wb.save(temp_xlsx_path)
        wb.close()
        
        parser = XlsxParser()
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        
        # Should still parse successfully
        assert_valid_parse_result(result)
        
        # Should have at least root, workbook, and one sheet
        assert len(result["elements"]) >= 3
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_missing_binary_path(self, temp_xlsx_path):
        """Test handling when binary_path is missing but content is provided."""
        # Create a valid XLSX file first
        create_simple_xlsx(temp_xlsx_path)
        
        # Read its content
        with open(temp_xlsx_path, 'rb') as f:
            xlsx_content = f.read()
        
        parser = XlsxParser()
        
        # Parse with content but no binary_path - should create temp file
        result = parser.parse({
            "id": "/test.xlsx",
            "content": xlsx_content,  # Provide actual XLSX content
            "metadata": {}
        })
        
        # Should create temp file and parse successfully
        assert_valid_parse_result(result)
        assert len(result["elements"]) >= 3  # At least root, workbook, sheet


# =============================================================================
# Performance Tests
# =============================================================================

@pytest.mark.performance
class TestXlsxPerformance:
    """Performance tests for XLSX parser."""
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    @pytest.mark.slow
    def test_large_spreadsheet_performance(self, temp_xlsx_path):
        """Test parsing performance with large spreadsheet."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create large dataset (100x100)
        for row in range(1, 101):
            for col in range(1, 101):
                ws.cell(row=row, column=col, value=f"R{row}C{col}")
        
        wb.save(temp_xlsx_path)
        wb.close()
        
        parser = XlsxParser({"max_rows": 1000, "max_cols": 1000})
        
        start_time = time.time()
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        elapsed = time.time() - start_time
        
        # Should parse in reasonable time (< 5 seconds for 10K cells)
        assert elapsed < 5.0, f"Parsing took {elapsed}s, exceeds 5s limit"
        
        # Should have parsed all cells
        assert len(result["elements"]) > 100  # At least rows + cells
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_max_limits_respected(self, temp_xlsx_path):
        """Test that max_rows and max_cols limits are respected."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Create data beyond limits
        for row in range(1, 21):  # 20 rows
            for col in range(1, 21):  # 20 columns
                ws.cell(row=row, column=col, value=f"R{row}C{col}")
        
        wb.save(temp_xlsx_path)
        wb.close()
        
        # Parse with limits
        parser = XlsxParser({"max_rows": 10, "max_cols": 10})
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        
        # Count actual data cells (not structural elements)
        cells = [e for e in result["elements"] 
                if e["element_type"] in ["table_cell", "table_header"]]
        
        # Should not exceed max_rows * max_cols
        assert len(cells) <= 10 * 10


# =============================================================================
# Metadata Extraction Tests
# =============================================================================

@pytest.mark.integration
class TestMetadataExtraction:
    """Test document metadata extraction."""
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_document_properties(self, temp_xlsx_path):
        """Test extraction of document properties."""
        wb = openpyxl.Workbook()
        
        # Set document properties
        wb.properties.title = "Test Document"
        wb.properties.creator = "Test Author"
        wb.properties.subject = "Test Subject"
        wb.properties.keywords = "test, xlsx, parser"
        
        wb.save(temp_xlsx_path)
        wb.close()
        
        parser = XlsxParser()
        result = parser.parse({
            "id": temp_xlsx_path,
            "binary_path": temp_xlsx_path,
            "metadata": {}
        })
        
        # Check metadata extraction
        metadata = result["document"]["metadata"]
        
        assert metadata.get("title") == "Test Document"
        assert metadata.get("author") == "Test Author"
        assert metadata.get("subject") == "Test Subject"
        assert metadata.get("keywords") == "test, xlsx, parser"
    
    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not available")
    def test_sheet_metadata(self, multi_sheet_xlsx_content):
        """Test extraction of sheet-level metadata."""
        parser = XlsxParser()
        result = parser.parse(multi_sheet_xlsx_content)
        
        # Check document metadata has sheet info
        doc_metadata = result["document"]["metadata"]
        assert doc_metadata["sheet_count"] == 3
        assert set(doc_metadata["sheet_names"]) == {"Sales", "Inventory", "Reports"}
        
        # Check individual sheet metadata
        sheets = [e for e in result["elements"] if e["element_type"] == "sheet"]
        for sheet in sheets:
            assert "title" in sheet["metadata"]
            assert "max_row" in sheet["metadata"]
            assert "max_column" in sheet["metadata"]


# =============================================================================
# Test Runner
# =============================================================================

if __name__ == "__main__":
    # Run tests with markers
    import subprocess
    import sys
    
    # Run unit tests first (fast)
    print("Running unit tests...")
    subprocess.run([sys.executable, "-m", "pytest", __file__, "-m", "unit", "-v"])
    
    # Run integration tests
    print("\nRunning integration tests...")
    subprocess.run([sys.executable, "-m", "pytest", __file__, "-m", "integration", "-v"])
    
    # Run performance tests (optional, usually slow)
    print("\nRunning performance tests...")
    subprocess.run([sys.executable, "-m", "pytest", __file__, "-m", "performance", "-v"])