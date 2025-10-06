#!/usr/bin/env python3
"""
Analyze the structure of our test XLSX file to understand merged cells.
"""

import openpyxl
from pathlib import Path

def analyze_xlsx():
    """Analyze the test XLSX file structure."""

    test_file = Path("tests/assets/test_sample.xlsx")
    if not test_file.exists():
        print("Test file not found")
        return

    print(f"Analyzing: {test_file}")
    print("=" * 50)

    # Load workbook
    wb = openpyxl.load_workbook(test_file)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        print(f"Max row: {sheet.max_row}, Max col: {sheet.max_column}")

        # Check for merged cells
        if hasattr(sheet, 'merged_cells') and sheet.merged_cells:
            print(f"Merged cell ranges: {len(sheet.merged_cells.ranges)}")
            for i, merged_range in enumerate(sheet.merged_cells.ranges):
                print(f"  Range {i+1}: {merged_range}")
                # Get the content of the merged cell (top-left cell)
                min_row, min_col = merged_range.min_row, merged_range.min_col
                max_row, max_col = merged_range.max_row, merged_range.max_col
                cell = sheet.cell(row=min_row, column=min_col)
                print(f"    Content: '{cell.value}'")
                print(f"    Spans: {max_row - min_row + 1} rows x {max_col - min_col + 1} cols")
        else:
            print("No merged cells found")

        # Show some sample cell content
        print("Sample cells:")
        for row in range(1, min(6, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(7, sheet.max_column + 1)):
                cell = sheet.cell(row=row, column=col)
                value = str(cell.value) if cell.value is not None else ""
                row_data.append(value[:10])  # Truncate for display
            print(f"  Row {row}: {row_data}")

if __name__ == "__main__":
    analyze_xlsx()